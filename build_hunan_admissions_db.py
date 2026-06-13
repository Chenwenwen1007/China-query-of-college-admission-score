from __future__ import annotations

import argparse
import concurrent.futures
import io
import json
import re
import sqlite3
import threading
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import pandas as pd
import pdfplumber
import requests
from openpyxl import load_workbook

from province_config import DEFAULT_PROVINCE_SLUG, load_province_config


BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "raw"


ACTIVE_PROVINCE_SLUG = DEFAULT_PROVINCE_SLUG
PROVINCE_CONFIG: dict[str, Any] = {}
DB_PATH = BASE_DIR / "hunan_gaokao_admissions_2023_2025.sqlite3"
PROVINCE_ID = ""
USER_SUBJECTS: set[str] = set()
SCORE_MIN = 0
SCORE_MAX = 750
SUBJECT_FILTER_TEXT = ""
PLAN_LOCAL_BATCH_NAMES: set[str] = set()
PLAN_TYPE_CODES: set[str] = set()
RAW_SOURCE_OPTIONS: dict[str, Any] = {}
OFFICIAL_SOURCES: dict[int, dict[str, Any]] = {}
LEGACY_SCHOOL_LINE_SOURCES: list[dict[str, Any]] = []

GAOKAO_SCHOOL_LIST_URL = "https://static-data.gaokao.cn/www/2.0/school/list_v2.json"
GAOKAO_SCHOOL_LIST_PATH = RAW_DIR / "gaokao_school_list_v2.json"
GAOKAO_SPECIAL_PLAN_URL = ""


def configure_runtime(province_slug: str) -> None:
    global ACTIVE_PROVINCE_SLUG
    global PROVINCE_CONFIG
    global DB_PATH
    global PROVINCE_ID
    global USER_SUBJECTS
    global SCORE_MIN
    global SCORE_MAX
    global SUBJECT_FILTER_TEXT
    global PLAN_LOCAL_BATCH_NAMES
    global PLAN_TYPE_CODES
    global RAW_SOURCE_OPTIONS
    global OFFICIAL_SOURCES
    global LEGACY_SCHOOL_LINE_SOURCES
    global GAOKAO_SPECIAL_PLAN_URL

    ACTIVE_PROVINCE_SLUG = province_slug
    PROVINCE_CONFIG = load_province_config(province_slug)
    DB_PATH = BASE_DIR / PROVINCE_CONFIG["database_filename"]
    PROVINCE_ID = str(PROVINCE_CONFIG["province_code"])
    USER_SUBJECTS = set(PROVINCE_CONFIG["user_subjects"])
    SCORE_MIN = int(PROVINCE_CONFIG["score_range"]["min"])
    SCORE_MAX = int(PROVINCE_CONFIG["score_range"]["max"])
    def clean_config_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    SUBJECT_FILTER_TEXT = clean_config_text(PROVINCE_CONFIG.get("score_filter", {}).get("subject_contains"))
    PLAN_LOCAL_BATCH_NAMES = {
        clean_config_text(value)
        for value in PROVINCE_CONFIG.get("plan_filter", {}).get("local_batch_names", [])
        if clean_config_text(value)
    }
    PLAN_TYPE_CODES = {
        clean_config_text(value)
        for value in PROVINCE_CONFIG.get("plan_filter", {}).get("type_codes", [])
        if clean_config_text(value)
    }
    RAW_SOURCE_OPTIONS = PROVINCE_CONFIG.get("raw_source_options", {})
    OFFICIAL_SOURCES = {}
    for year, payload in PROVINCE_CONFIG["official_sources"].items():
        source_payload = dict(payload)
        source_payload["local_path"] = RAW_DIR / payload["local_filename"]
        source_payload["source_format"] = payload.get("source_format", "xlsx")
        OFFICIAL_SOURCES[int(year)] = source_payload
    LEGACY_SCHOOL_LINE_SOURCES = []
    for source_key, payload in PROVINCE_CONFIG.get("legacy_school_line_sources", {}).items():
        LEGACY_SCHOOL_LINE_SOURCES.append(
            {
                "source_key": source_key,
                "year": int(payload["year"]),
                "batch": clean_config_text(payload.get("batch")),
                "title": payload["title"],
                "publisher": payload["publisher"],
                "landing_url": payload["landing_url"],
                "file_url": payload["file_url"],
                "local_path": RAW_DIR / payload["local_filename"],
                "source_format": payload.get("source_format", "pdf"),
                "subject_contains": clean_config_text(payload.get("subject_contains")),
            }
        )
    GAOKAO_SPECIAL_PLAN_URL = (
        f"https://static-data.gaokao.cn/www/2.0/schoolspecialplan/{{school_id}}/{{year}}/{PROVINCE_ID}.json"
    )


configure_runtime(DEFAULT_PROVINCE_SLUG)


def validate_runtime_build_ready() -> None:
    if not PROVINCE_CONFIG.get("build_enabled", True):
        blockers = PROVINCE_CONFIG.get("build_blockers", [])
        details = "\n".join(f"- {item}" for item in blockers)
        raise RuntimeError(
            f"province '{ACTIVE_PROVINCE_SLUG}' is not build-ready yet.\n{details}".rstrip()
        )

    missing_fields: list[str] = []
    for year, source in PROVINCE_CONFIG.get("official_sources", {}).items():
        for field in ("landing_url", "file_url", "local_filename"):
            if not clean_text(source.get(field)):
                missing_fields.append(f"official_sources[{year}].{field}")
    if missing_fields:
        raise RuntimeError(
            f"province '{ACTIVE_PROVINCE_SLUG}' has incomplete official source config: {', '.join(missing_fields)}"
        )

MANUAL_SCHOOL_ALIASES = {
    "云南大学滇池学院": "滇池学院",
    "佛山科学技术学院": "佛山大学",
    "合肥学院": "合肥大学",
    "嘉兴学院": "嘉兴大学",
    "天水师范学院": "天水师范大学",
    "宁夏师范学院": "宁夏师范大学",
    "浙江科技学院": "浙江科技大学",
    "海南医学院": "海南医科大学",
    "湖州师范学院": "湖州师范大学",
    "潍坊医学院": "山东第二医科大学",
    "牡丹江医学院": "牡丹江医科大学",
    "蚌埠医学院": "蚌埠医科大学",
    "西藏农牧学院": "西藏农牧大学",
    "滨州学院": "山东航空学院",
}

MANUAL_SCHOOL_ID_OVERRIDES = {
    "华北科技学院": ("604", "华北科技学院"),
    "吉林化工学院": ("400", "吉林化工学院"),
    "吉首大学张家界学院": ("2476", "吉首大学张家界学院"),
    "四川外国语大学成都学院": ("2497", "四川外国语大学成都学院"),
    "安徽科技学院": ("331", "安徽科技学院"),
    "常熟理工学院": ("1182", "常熟理工学院"),
    "榆林学院": ("1090", "榆林学院"),
    "武汉城市学院（原武汉科技大学城市学院）": ("1176", "武汉城市学院"),
    "淮阴工学院": ("152", "淮阴工学院"),
    "湖南师范大学树达学院": ("2466", "湖南师范大学树达学院"),
    "湖南文理学院芙蓉学院": ("2474", "湖南文理学院芙蓉学院"),
    "湖南理工学院": ("396", "湖南理工学院"),
    "湖南理工学院南湖学院": ("2475", "湖南理工学院南湖学院"),
    "湘潭大学兴湘学院": ("2472", "湘潭大学兴湘学院"),
    "绍兴文理学院": ("255", "绍兴文理学院"),
    "赤峰学院": ("1042", "赤峰学院"),
    "重庆三峡学院": ("198", "重庆三峡学院"),
    "闽江学院": ("477", "闽江学院"),
    "防灾科技学院": ("653", "防灾科技学院"),
    "青海大学昆仑学院": ("1555", "青海大学昆仑学院"),
    "陆军军医大学": ("1227", "中国人民解放军陆军军医大学"),
    "空军军医大学": ("937", "中国人民解放军空军军医大学"),
    "国防科技大学": ("939", "中国人民解放军国防科技大学"),
    "西南大学(荣昌校区)": ("934", "西南大学"),
    "山东大学威海分校": ("1219", "山东大学（威海）"),
    "电子科技大学(沙河校区)": ("661", "电子科技大学"),
    "复旦大学医学院": ("3427", "复旦大学上海医学院"),
    "桂林医学院": ("538", "桂林医科大学"),
    "赣南医学院": ("195", "赣南医科大学"),
    "重庆科技学院": ("1009", "重庆科技大学"),
    "北京邮电大学(宏福校区)": ("48", "北京邮电大学"),
    "北京师范大学-香港浸会大学联合国际学院": ("1275", "北师香港浸会大学"),
    "云南艺术学院文华学院": ("2525", "昆明传媒学院"),
    "南昌工程学院": ("867", "江西水利电力大学"),
    "新乡医学院": ("482", "河南医药大学"),
    "滨州医学院": ("64", "山东医药大学"),
}


COL_BATCH = "批次"
COL_PLAN_CATEGORY = "计划类别"
COL_SUBJECT = "科类"
COL_SCHOOL_CODE = "院校代号"
COL_SCHOOL_NAME = "院校名称"
COL_GROUP_CODE = "专业组编号"
COL_GROUP_NAME = "专业组名称"
COL_SCORE = "投档线"
COL_NOTE = "备注"
COL_MATH_LANG_SUM = "语数之和"
COL_MATH_LANG_MAX = "语数最高"
COL_FOREIGN = "外语"
COL_FIRST_CHOICE = "首选科目"
COL_SECOND_MAX = "再选最高"
COL_SECOND_NEXT = "再选次高"
COL_VOLUNTEER_ORDER = "志愿序号"


@dataclass(frozen=True)
class ScoreRow:
    year: int
    batch: str
    plan_category: str
    subject_type: str
    school_code: str
    school_name: str
    group_code: str
    group_name: str
    score: int
    note: str
    math_lang_sum: str
    math_lang_max: str
    foreign_score: str
    first_choice_score: str
    second_max: str
    second_next: str
    volunteer_order: str
    source_url: str


@dataclass(frozen=True)
class LegacySchoolLineRow:
    year: int
    batch: str
    subject_type: str
    school_code: str
    school_name: str
    score: int
    rank_value: str
    note: str
    source_url: str


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def ensure_dirs() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)


def download_file(session: requests.Session, url: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.stat().st_size > 0:
        return
    response = session.get(url, timeout=60)
    response.raise_for_status()
    path.write_bytes(response.content)


def ensure_raw_sources(session: requests.Session) -> Path:
    ensure_dirs()
    for source in OFFICIAL_SOURCES.values():
        download_file(session, source["file_url"], source["local_path"])
    for source in LEGACY_SCHOOL_LINE_SOURCES:
        download_file(session, source["file_url"], source["local_path"])
    zip_source_year = RAW_SOURCE_OPTIONS.get("zip_source_year")
    extract_dir = RAW_DIR / str(RAW_SOURCE_OPTIONS.get("zip_extract_dir", "zip_extract"))
    if zip_source_year is None:
        return extract_dir
    zip_path = OFFICIAL_SOURCES[int(zip_source_year)]["local_path"]
    if not extract_dir.exists():
        extract_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(zip_path) as zf:
            zf.extractall(extract_dir)
    return extract_dir


def ensure_gaokao_school_list(session: requests.Session) -> None:
    download_file(session, GAOKAO_SCHOOL_LIST_URL, GAOKAO_SCHOOL_LIST_PATH)


def load_zip_score_source(extract_dir: Path) -> pd.DataFrame:
    zip_glob = str(RAW_SOURCE_OPTIONS.get("zip_filename_glob", "*.xls"))
    xls_paths = sorted(extract_dir.glob(zip_glob))
    if not xls_paths:
        raise FileNotFoundError(f"zip score source not found after unzip: {extract_dir / zip_glob}")
    header_row = int(RAW_SOURCE_OPTIONS.get("zip_header_row", 1))
    df = pd.read_excel(xls_paths[0], header=header_row)
    df.columns = [str(col).strip() for col in df.columns]
    zip_plan_category = clean_text(RAW_SOURCE_OPTIONS.get("zip_plan_category_label"))
    zip_batch = clean_text(RAW_SOURCE_OPTIONS.get("zip_batch_label"))
    if zip_plan_category:
        df.insert(0, COL_PLAN_CATEGORY, zip_plan_category)
    if zip_batch:
        df.insert(0, COL_BATCH, zip_batch)
    return df


def load_xlsx_scores(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_row_index = int(RAW_SOURCE_OPTIONS.get("xlsx_header_row_index", 2))
    data_start_row_index = int(RAW_SOURCE_OPTIONS.get("xlsx_data_start_row_index", 3))
    header = [str(value).strip() if value is not None else "" for value in rows[header_row_index]]
    data = rows[data_start_row_index:]
    return pd.DataFrame(data, columns=header)


def decode_zip_member_name(member_name: str) -> str:
    try:
        return member_name.encode("cp437").decode("gbk")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return member_name


def select_zip_member_by_keywords(zf: zipfile.ZipFile, keywords: list[str]) -> zipfile.ZipInfo:
    normalized_keywords = [clean_text(keyword) for keyword in keywords if clean_text(keyword)]
    if not normalized_keywords:
        raise ValueError("zip member keyword list is empty")

    for info in zf.infolist():
        decoded_name = decode_zip_member_name(info.filename)
        if all(keyword in decoded_name for keyword in normalized_keywords):
            return info
    raise FileNotFoundError(f"zip member not found for keywords: {normalized_keywords}")


JX_WATERMARK_CHARS = set("\u6c5f\u897f\u7701\u6559\u80b2\u8003\u8bd5\u9662")
JX_PDF_SUBJECT_LABELS = (
    "\u5386\u53f2\u7c7b",
    "\u7269\u7406\u7c7b",
    "\u4e09\u6821\u751f\u7c7b",
    "\u6587\u53f2",
    "\u7406\u5de5",
    "\u4e09\u6821\u751f\u6587\u7406\u7c7b",
    "\u4e09\u6821\u751f\u6587\u7406",
)


def strip_single_watermark_prefix(text: str) -> str:
    value = text.strip()
    while len(value) >= 2 and value[0] in JX_WATERMARK_CHARS and value[1] not in JX_WATERMARK_CHARS:
        value = value[1:]
    return value


def join_pdf_cell_parts(value: Any) -> str:
    if value is None:
        return ""
    parts = [part.strip() for part in str(value).replace("\r", "\n").split("\n") if part.strip()]
    return "".join(parts)


def extract_subject_label(value: str) -> str:
    text = strip_single_watermark_prefix(value)
    for label in JX_PDF_SUBJECT_LABELS:
        if label in text:
            return label
    return text


def extract_group_name(value: str) -> str:
    text = strip_single_watermark_prefix(value).replace("\uff08", "(").replace("\uff09", ")")
    marker = "\u7b2c"
    index = text.find(marker)
    return text[index:] if index >= 0 else text


def detect_jx_legacy_batch(page_text: str) -> str:
    first_batch = "\u7b2c\u4e00\u6279\u672c\u79d1"
    second_batch = "\u7b2c\u4e8c\u6279\u672c\u79d1"
    if first_batch in page_text:
        return "\u672c\u79d1\u4e00\u6279"
    if second_batch in page_text:
        return "\u672c\u79d1\u4e8c\u6279"
    return ""


def load_jiangxi_pdf_scores(path: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            page_batch = detect_jx_legacy_batch(page_text)
            for table in page.extract_tables() or []:
                for raw_row in table[1:]:
                    if not raw_row:
                        continue
                    cells = [join_pdf_cell_parts(cell) for cell in raw_row]
                    if len(cells) >= 8:
                        subject_type = extract_subject_label(cells[1])
                        school_code = clean_text(re.search(r"\d+", strip_single_watermark_prefix(cells[2])).group(0) if re.search(r"\d+", strip_single_watermark_prefix(cells[2])) else "")
                        school_name = join_pdf_cell_parts(raw_row[3])
                        group_code = clean_text(re.search(r"\d+", strip_single_watermark_prefix(cells[4])).group(0) if re.search(r"\d+", strip_single_watermark_prefix(cells[4])) else "")
                        group_name = extract_group_name(cells[5])
                        score_match = re.search(r"\d+", strip_single_watermark_prefix(cells[6]))
                        rank_match = re.search(r"\d+", strip_single_watermark_prefix(cells[7]))
                        if subject_type not in JX_PDF_SUBJECT_LABELS or not school_code or not school_name or score_match is None:
                            continue
                        rows.append(
                            {
                                COL_BATCH: page_batch or "\u672c\u79d1\u6279",
                                COL_PLAN_CATEGORY: "",
                                COL_SUBJECT: subject_type,
                                COL_SCHOOL_CODE: school_code,
                                COL_SCHOOL_NAME: school_name,
                                COL_GROUP_CODE: group_code,
                                COL_GROUP_NAME: group_name,
                                COL_SCORE: score_match.group(0),
                                COL_NOTE: clean_text(cells[8]) if len(cells) > 8 else "",
                                COL_MATH_LANG_SUM: "",
                                COL_MATH_LANG_MAX: "",
                                COL_FOREIGN: "",
                                COL_FIRST_CHOICE: "",
                                COL_SECOND_MAX: "",
                                COL_SECOND_NEXT: "",
                                COL_VOLUNTEER_ORDER: rank_match.group(0) if rank_match else "",
                            }
                        )
                    elif len(cells) >= 6:
                        subject_type = extract_subject_label(cells[1])
                        score_match = re.search(r"\d+", cells[4])
                        rank_match = re.search(r"\d+", cells[5])
                        if subject_type not in JX_PDF_SUBJECT_LABELS or score_match is None:
                            continue
                        rows.append(
                            {
                                COL_BATCH: page_batch,
                                COL_PLAN_CATEGORY: "",
                                COL_SUBJECT: subject_type,
                                COL_SCHOOL_CODE: clean_text(cells[2]),
                                COL_SCHOOL_NAME: clean_text(cells[3]),
                                COL_GROUP_CODE: "",
                                COL_GROUP_NAME: "",
                                COL_SCORE: score_match.group(0),
                                COL_NOTE: "",
                                COL_MATH_LANG_SUM: "",
                                COL_MATH_LANG_MAX: "",
                                COL_FOREIGN: "",
                                COL_FIRST_CHOICE: "",
                                COL_SECOND_MAX: "",
                                COL_SECOND_NEXT: "",
                                COL_VOLUNTEER_ORDER: rank_match.group(0) if rank_match else "",
                            }
                        )
    if not rows:
        raise RuntimeError(f"no score rows parsed from Jiangxi PDF: {path}")
    return pd.DataFrame(rows)


def load_guangdong_pdf_scores_from_bytes(pdf_bytes: bytes) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for raw_row in table:
                    if not raw_row:
                        continue
                    cells = [join_pdf_cell_parts(cell) for cell in raw_row]
                    if len(cells) < 7:
                        continue

                    if "院校代码" in cells[0] and "院校名称" in cells[1]:
                        continue

                    school_code_match = re.search(r"\d+", cells[0])
                    group_code_match = re.search(r"\d+", cells[2])
                    score_match = re.search(r"\d+", cells[5])
                    rank_match = re.search(r"\d+", cells[6])
                    if school_code_match is None or group_code_match is None or score_match is None:
                        continue

                    school_code = school_code_match.group(0)
                    school_name = clean_text(cells[1])
                    group_code = group_code_match.group(0)
                    note_parts: list[str] = []
                    if clean_text(cells[2]) != group_code:
                        note_parts.append(f"原专业组代码: {clean_text(cells[2])}")
                    if clean_text(cells[5]) != score_match.group(0):
                        note_parts.append(f"原投档最低分: {clean_text(cells[5])}")

                    rows.append(
                        {
                            COL_BATCH: "本科批",
                            COL_PLAN_CATEGORY: "普通类",
                            COL_SUBJECT: "物理类",
                            COL_SCHOOL_CODE: school_code,
                            COL_SCHOOL_NAME: school_name,
                            COL_GROUP_CODE: group_code,
                            COL_GROUP_NAME: f"第{group_code}专业组",
                            COL_SCORE: score_match.group(0),
                            COL_NOTE: "；".join(note_parts),
                            COL_MATH_LANG_SUM: "",
                            COL_MATH_LANG_MAX: "",
                            COL_FOREIGN: "",
                            COL_FIRST_CHOICE: "",
                            COL_SECOND_MAX: "",
                            COL_SECOND_NEXT: "",
                            COL_VOLUNTEER_ORDER: rank_match.group(0) if rank_match else "",
                        }
                    )

    if not rows:
        raise RuntimeError("no score rows parsed from Guangdong PDF")
    return pd.DataFrame(rows)


def load_guangdong_pdf_scores(path: Path) -> pd.DataFrame:
    return load_guangdong_pdf_scores_from_bytes(path.read_bytes())


def load_guangdong_zip_pdf_scores(path: Path, source: dict[str, Any]) -> pd.DataFrame:
    keywords = [clean_text(value) for value in source.get("zip_member_keywords", []) if clean_text(value)]
    with zipfile.ZipFile(path) as zf:
        target = select_zip_member_by_keywords(zf, keywords)
        return load_guangdong_pdf_scores_from_bytes(zf.read(target))


def load_scores_for_source(source: dict[str, Any], extract_dir: Path) -> pd.DataFrame:
    source_format = clean_text(source.get("source_format")) or "xlsx"
    zip_source_year = RAW_SOURCE_OPTIONS.get("zip_source_year")
    if zip_source_year is not None and int(source["year"]) == int(zip_source_year):
        return load_zip_score_source(extract_dir)
    if source_format == "xlsx":
        return load_xlsx_scores(source["local_path"])
    if source_format == "pdf":
        if ACTIVE_PROVINCE_SLUG == "jiangxi":
            return load_jiangxi_pdf_scores(source["local_path"])
        if ACTIVE_PROVINCE_SLUG == "guangdong":
            return load_guangdong_pdf_scores(source["local_path"])
        raise NotImplementedError(f"pdf source parser not implemented for province '{ACTIVE_PROVINCE_SLUG}'")
    if source_format == "zip_pdf":
        if ACTIVE_PROVINCE_SLUG == "guangdong":
            return load_guangdong_zip_pdf_scores(source["local_path"], source)
        raise NotImplementedError(f"zip_pdf source parser not implemented for province '{ACTIVE_PROVINCE_SLUG}'")
    raise ValueError(f"unsupported source_format: {source_format}")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def numeric_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def load_filtered_scores(extract_dir: Path) -> list[ScoreRow]:
    files: dict[int, pd.DataFrame] = {}
    for year, source in OFFICIAL_SOURCES.items():
        source_with_year = dict(source)
        source_with_year["year"] = year
        files[year] = load_scores_for_source(source_with_year, extract_dir)

    score_rows: list[ScoreRow] = []
    for year, df in files.items():
        frame = df.copy()
        frame[COL_SCORE] = pd.to_numeric(frame[COL_SCORE], errors="coerce")
        if SUBJECT_FILTER_TEXT:
            frame = frame[frame[COL_SUBJECT].astype(str).str.contains(SUBJECT_FILTER_TEXT, na=False)]
        frame = frame[(frame[COL_SCORE] >= SCORE_MIN) & (frame[COL_SCORE] <= SCORE_MAX)]
        for _, row in frame.iterrows():
            score = numeric_int(row.get(COL_SCORE))
            if score is None:
                continue
            score_rows.append(
                ScoreRow(
                    year=year,
                    batch=clean_text(row.get(COL_BATCH)),
                    plan_category=clean_text(row.get(COL_PLAN_CATEGORY)),
                    subject_type=clean_text(row.get(COL_SUBJECT)),
                    school_code=clean_text(row.get(COL_SCHOOL_CODE)),
                    school_name=clean_text(row.get(COL_SCHOOL_NAME)),
                    group_code=clean_text(row.get(COL_GROUP_CODE)),
                    group_name=clean_text(row.get(COL_GROUP_NAME)),
                    score=score,
                    note=clean_text(row.get(COL_NOTE)),
                    math_lang_sum=clean_text(row.get(COL_MATH_LANG_SUM)),
                    math_lang_max=clean_text(row.get(COL_MATH_LANG_MAX)),
                    foreign_score=clean_text(row.get(COL_FOREIGN)),
                    first_choice_score=clean_text(row.get(COL_FIRST_CHOICE)),
                    second_max=clean_text(row.get(COL_SECOND_MAX)),
                    second_next=clean_text(row.get(COL_SECOND_NEXT)),
                    volunteer_order=clean_text(row.get(COL_VOLUNTEER_ORDER)),
                    source_url=OFFICIAL_SOURCES[year]["landing_url"],
                )
            )
    return score_rows


def load_legacy_school_line_rows() -> list[LegacySchoolLineRow]:
    rows: list[LegacySchoolLineRow] = []
    for source in LEGACY_SCHOOL_LINE_SOURCES:
        frame = load_scores_for_source(source, RAW_DIR)
        subject_filter = clean_text(source.get("subject_contains"))
        if subject_filter:
            frame = frame[frame[COL_SUBJECT].astype(str).str.contains(subject_filter, na=False)]
        frame[COL_SCORE] = pd.to_numeric(frame[COL_SCORE], errors="coerce")
        frame = frame[(frame[COL_SCORE] >= SCORE_MIN) & (frame[COL_SCORE] <= SCORE_MAX)]
        for _, row in frame.iterrows():
            school_code = clean_text(row.get(COL_SCHOOL_CODE))
            school_name = clean_text(row.get(COL_SCHOOL_NAME))
            score = numeric_int(row.get(COL_SCORE))
            if not school_code or not school_name or score is None:
                continue
            rows.append(
                LegacySchoolLineRow(
                    year=int(source["year"]),
                    batch=clean_text(row.get(COL_BATCH)) or clean_text(source.get("batch")),
                    subject_type=clean_text(row.get(COL_SUBJECT)),
                    school_code=school_code,
                    school_name=school_name,
                    score=score,
                    rank_value=clean_text(row.get(COL_VOLUNTEER_ORDER)),
                    note=clean_text(row.get(COL_NOTE)),
                    source_url=source["landing_url"],
                )
            )
    unique_rows: dict[tuple[int, str, str, str], LegacySchoolLineRow] = {}
    for row in rows:
        unique_rows[(row.year, row.batch, row.school_code, row.school_name)] = row
    return list(unique_rows.values())


def normalize_school_name(name: str) -> str:
    value = clean_text(name)
    value = value.replace("（", "(").replace("）", ")").replace("·", "")
    return re.sub(r"\s+", "", value)


def build_school_name_variants(name: str) -> list[str]:
    variants: list[str] = []

    def add_variant(value: str) -> None:
        normalized = normalize_school_name(value)
        if normalized and normalized not in variants:
            variants.append(normalized)

    base_name = clean_text(name)
    add_variant(base_name)

    if ACTIVE_PROVINCE_SLUG == "guangdong":
        stripped = re.sub(
            r"\((?:地方专项|中外合作办学|联合培养|校企联合培养|学分互认|协同培养|国际班|汕尾校区|揭阳校区|河源校区|荣昌校区|威海校区|宏福校区|马来西亚分校|闽台合作|原[^)]*)\)",
            "",
            base_name,
        )
        add_variant(stripped)
        add_variant(re.sub(r"\([^)]*\)", "", base_name))

        if "佛山大学" in base_name:
            add_variant(base_name.replace("佛山大学", "佛山科学技术学院"))
            add_variant(re.sub(r"\([^)]*\)", "", base_name).replace("佛山大学", "佛山科学技术学院"))

    return variants


def extract_digits(value: str) -> str:
    return "".join(re.findall(r"\d+", clean_text(value)))


def normalize_group_name(value: str) -> str:
    text = clean_text(value)
    text = text.replace("（", "(").replace("）", ")")
    return text


def official_group_category(group_name: str) -> str:
    text = clean_text(group_name)
    if "国家专项" in text:
        return "国家专项"
    if "中外合作办学" in text or "国际项目" in text:
        return "中外合作办学或国际项目"
    if "地方专项" in text:
        return "地方专项"
    if "民族班" in text or "少数民族" in text:
        return "民族计划"
    return "普通类"


def load_gaokao_school_map() -> dict[str, Any]:
    obj = json.loads(GAOKAO_SCHOOL_LIST_PATH.read_text(encoding="utf-8"))
    if obj.get("code") != "0000":
        raise RuntimeError("failed to load gaokao school list")
    return obj["data"]


def build_school_lookup(gaokao_data: dict[str, Any]) -> tuple[dict[str, tuple[str, dict[str, Any], str]], dict[str, list[tuple[str, dict[str, Any]]]]]:
    exact: dict[str, tuple[str, dict[str, Any], str]] = {}
    normalized: dict[str, list[tuple[str, dict[str, Any]]]] = {}
    for school_id, payload in gaokao_data.items():
        school_name = payload["name"]
        exact[school_name] = (school_id, payload, "exact")
        for variant in build_school_name_variants(school_name):
            normalized.setdefault(variant, []).append((school_id, payload))
    return exact, normalized


def match_school_name(
    school_name: str,
    exact_lookup: dict[str, tuple[str, dict[str, Any], str]],
    normalized_lookup: dict[str, list[tuple[str, dict[str, Any]]]],
) -> tuple[str | None, str | None, str, float, str]:
    cleaned_variants = [school_name]
    cleaned_variant = strip_single_watermark_prefix(school_name)
    if cleaned_variant != school_name:
        cleaned_variants.append(cleaned_variant)
    if ACTIVE_PROVINCE_SLUG == "jiangxi":
        for prefix_length in range(1, 4):
            if len(school_name) > prefix_length and all(char in JX_WATERMARK_CHARS for char in school_name[:prefix_length]):
                candidate = school_name[prefix_length:]
                if candidate not in cleaned_variants:
                    cleaned_variants.append(candidate)

    manual_override = MANUAL_SCHOOL_ID_OVERRIDES.get(school_name)
    if manual_override:
        school_id, gaokao_school_name = manual_override
        return school_id, gaokao_school_name, "manual_school_id", 0.99, f"https://www.gaokao.cn/school/{school_id}"

    if school_name in exact_lookup:
        school_id, payload, match_type = exact_lookup[school_name]
        return school_id, payload["name"], match_type, 1.0, ""

    alias_name = MANUAL_SCHOOL_ALIASES.get(school_name)
    if alias_name and alias_name in exact_lookup:
        school_id, payload, _ = exact_lookup[alias_name]
        return school_id, payload["name"], "manual_alias", 0.95, school_name

    for candidate_name in cleaned_variants[1:]:
        if candidate_name in exact_lookup:
            school_id, payload, _ = exact_lookup[candidate_name]
            return school_id, payload["name"], "watermark_prefix_cleaned", 0.92, school_name
        alias_name = MANUAL_SCHOOL_ALIASES.get(candidate_name)
        if alias_name and alias_name in exact_lookup:
            school_id, payload, _ = exact_lookup[alias_name]
            return school_id, payload["name"], "watermark_prefix_alias", 0.9, school_name

    for candidate_name in cleaned_variants:
        for variant in build_school_name_variants(candidate_name):
            candidates = normalized_lookup.get(variant, [])
            if len(candidates) == 1:
                school_id, payload = candidates[0]
                notes = "" if candidate_name == school_name else school_name
                match_type = "normalized" if candidate_name == school_name else "watermark_prefix_normalized"
                confidence = 0.9 if candidate_name == school_name else 0.88
                return school_id, payload["name"], match_type, confidence, notes
            if len(candidates) > 1:
                return None, None, "ambiguous", 0.0, json.dumps([payload["name"] for _, payload in candidates], ensure_ascii=False)
    return None, None, "unmatched", 0.0, ""


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode = WAL;
        PRAGMA synchronous = NORMAL;

        DROP TABLE IF EXISTS sources;
        DROP TABLE IF EXISTS crawl_runs;
        DROP TABLE IF EXISTS admission_groups;
        DROP TABLE IF EXISTS school_mappings;
        DROP TABLE IF EXISTS plan_major_details;
        DROP TABLE IF EXISTS eligible_majors;
        DROP TABLE IF EXISTS unmatched_groups;
        DROP TABLE IF EXISTS legacy_school_lines;

        CREATE TABLE sources (
            source_id INTEGER PRIMARY KEY,
            source_key TEXT NOT NULL UNIQUE,
            source_type TEXT NOT NULL,
            year INTEGER,
            title TEXT,
            publisher TEXT,
            landing_url TEXT,
            file_url TEXT,
            local_path TEXT,
            fetched_at TEXT NOT NULL
        );

        CREATE TABLE crawl_runs (
            run_id INTEGER PRIMARY KEY,
            started_at TEXT NOT NULL,
            finished_at TEXT,
            score_group_count INTEGER DEFAULT 0,
            school_count INTEGER DEFAULT 0,
            matched_school_count INTEGER DEFAULT 0,
            plan_fetch_success_count INTEGER DEFAULT 0,
            plan_fetch_failure_count INTEGER DEFAULT 0,
            eligible_major_count INTEGER DEFAULT 0,
            unmatched_group_count INTEGER DEFAULT 0,
            notes TEXT
        );

        CREATE TABLE admission_groups (
            admission_group_id INTEGER PRIMARY KEY,
            year INTEGER NOT NULL,
            batch TEXT NOT NULL,
            plan_category TEXT,
            subject_type TEXT NOT NULL,
            school_code TEXT NOT NULL,
            school_name TEXT NOT NULL,
            official_group_code TEXT NOT NULL,
            official_group_name TEXT NOT NULL,
            official_group_category TEXT NOT NULL,
            score INTEGER NOT NULL,
            note TEXT,
            math_lang_sum TEXT,
            math_lang_max TEXT,
            foreign_score TEXT,
            first_choice_score TEXT,
            second_max TEXT,
            second_next TEXT,
            volunteer_order TEXT,
            source_url TEXT NOT NULL
        );

        CREATE TABLE school_mappings (
            mapping_id INTEGER PRIMARY KEY,
            school_name_official TEXT NOT NULL,
            gaokao_school_id TEXT,
            gaokao_school_name TEXT,
            match_type TEXT NOT NULL,
            confidence REAL NOT NULL,
            alias_from TEXT,
            notes TEXT
        );

        CREATE TABLE plan_major_details (
            plan_major_id INTEGER PRIMARY KEY,
            year INTEGER NOT NULL,
            gaokao_school_id TEXT NOT NULL,
            gaokao_school_name TEXT NOT NULL,
            province_id TEXT NOT NULL,
            type_code TEXT NOT NULL,
            batch_code TEXT NOT NULL,
            local_batch_name TEXT,
            special_group_internal_id TEXT,
            special_group_display_name TEXT,
            group_digits TEXT,
            group_subject_info TEXT,
            group_subject_code TEXT,
            zslx_name TEXT,
            zslx_code TEXT,
            major_id TEXT,
            major_code TEXT,
            major_name TEXT,
            major_name_short TEXT,
            major_category_level1 TEXT,
            major_category_level2 TEXT,
            major_category_level3 TEXT,
            study_length TEXT,
            tuition TEXT,
            enrollment_count INTEGER,
            remark TEXT,
            extra_info TEXT,
            source_url TEXT NOT NULL,
            raw_payload TEXT NOT NULL
        );

        CREATE TABLE eligible_majors (
            eligible_major_id INTEGER PRIMARY KEY,
            admission_group_id INTEGER NOT NULL,
            plan_major_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            school_code TEXT NOT NULL,
            school_name TEXT NOT NULL,
            official_group_code TEXT NOT NULL,
            official_group_name TEXT NOT NULL,
            official_group_category TEXT NOT NULL,
            score INTEGER NOT NULL,
            elective_info TEXT,
            major_code TEXT,
            major_name TEXT NOT NULL,
            major_name_short TEXT,
            major_category_level1 TEXT,
            major_category_level2 TEXT,
            major_category_level3 TEXT,
            study_length TEXT,
            tuition TEXT,
            enrollment_count INTEGER,
            zslx_name TEXT,
            source_url TEXT NOT NULL,
            FOREIGN KEY(admission_group_id) REFERENCES admission_groups(admission_group_id),
            FOREIGN KEY(plan_major_id) REFERENCES plan_major_details(plan_major_id)
        );

        CREATE TABLE unmatched_groups (
            unmatched_group_id INTEGER PRIMARY KEY,
            admission_group_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            school_name TEXT NOT NULL,
            official_group_code TEXT NOT NULL,
            official_group_name TEXT NOT NULL,
            reason TEXT NOT NULL,
            detail TEXT,
            FOREIGN KEY(admission_group_id) REFERENCES admission_groups(admission_group_id)
        );

        CREATE TABLE legacy_school_lines (
            legacy_school_line_id INTEGER PRIMARY KEY,
            year INTEGER NOT NULL,
            batch TEXT NOT NULL,
            subject_type TEXT NOT NULL,
            school_code TEXT NOT NULL,
            school_name TEXT NOT NULL,
            score INTEGER NOT NULL,
            rank_value TEXT,
            note TEXT,
            source_url TEXT NOT NULL
        );

        CREATE INDEX idx_admission_groups_year_score ON admission_groups(year, score);
        CREATE INDEX idx_admission_groups_school ON admission_groups(year, school_name, official_group_code);
        CREATE INDEX idx_plan_major_school_year ON plan_major_details(year, gaokao_school_id, group_digits);
        CREATE INDEX idx_eligible_major_school_year ON eligible_majors(year, school_name);
        CREATE INDEX idx_legacy_school_lines_year_score ON legacy_school_lines(year, score);
        CREATE INDEX idx_legacy_school_lines_school ON legacy_school_lines(year, school_name);
        """
    )


def insert_sources(conn: sqlite3.Connection) -> None:
    fetched_at = now_iso()
    rows = [
        (
            f"official_scores_{year}",
            "official_scores",
            year,
            payload["title"],
            payload["publisher"],
            payload["landing_url"],
            payload["file_url"],
            str(payload["local_path"]),
            fetched_at,
        )
        for year, payload in OFFICIAL_SOURCES.items()
    ]
    for payload in LEGACY_SCHOOL_LINE_SOURCES:
        rows.append(
            (
                payload["source_key"],
                "legacy_school_line",
                payload["year"],
                payload["title"],
                payload["publisher"],
                payload["landing_url"],
                payload["file_url"],
                str(payload["local_path"]),
                fetched_at,
            )
        )
    rows.append(
        (
            "gaokao_school_list",
            "gaokao_school_list",
            None,
            "gaokao.cn 学校列表",
            "掌上高考",
            "https://m.gaokao.cn/",
            GAOKAO_SCHOOL_LIST_URL,
            str(GAOKAO_SCHOOL_LIST_PATH),
            fetched_at,
        )
    )
    conn.executemany(
        """
        INSERT INTO sources (
            source_key, source_type, year, title, publisher, landing_url, file_url, local_path, fetched_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )


def insert_legacy_school_lines(conn: sqlite3.Connection, rows: list[LegacySchoolLineRow]) -> int:
    if not rows:
        return 0
    conn.executemany(
        """
        INSERT INTO legacy_school_lines (
            year, batch, subject_type, school_code, school_name, score, rank_value, note, source_url
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row.year,
                row.batch,
                row.subject_type,
                row.school_code,
                row.school_name,
                row.score,
                row.rank_value,
                row.note,
                row.source_url,
            )
            for row in rows
        ],
    )
    return len(rows)


def insert_admission_groups(conn: sqlite3.Connection, score_rows: list[ScoreRow]) -> dict[tuple[int, str, str, str], int]:
    mapping: dict[tuple[int, str, str, str], int] = {}
    cursor = conn.cursor()
    for row in score_rows:
        cursor.execute(
            """
            INSERT INTO admission_groups (
                year, batch, plan_category, subject_type, school_code, school_name,
                official_group_code, official_group_name, official_group_category, score,
                note, math_lang_sum, math_lang_max, foreign_score, first_choice_score,
                second_max, second_next, volunteer_order, source_url
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row.year,
                row.batch,
                row.plan_category,
                row.subject_type,
                row.school_code,
                row.school_name,
                row.group_code,
                row.group_name,
                official_group_category(row.group_name),
                row.score,
                row.note,
                row.math_lang_sum,
                row.math_lang_max,
                row.foreign_score,
                row.first_choice_score,
                row.second_max,
                row.second_next,
                row.volunteer_order,
                row.source_url,
            ),
        )
        mapping[(row.year, row.school_name, row.group_code, row.group_name)] = cursor.lastrowid
    return mapping


def insert_school_mappings(
    conn: sqlite3.Connection,
    school_names: list[str],
    exact_lookup: dict[str, tuple[str, dict[str, Any], str]],
    normalized_lookup: dict[str, list[tuple[str, dict[str, Any]]]],
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    rows = []
    for school_name in sorted(set(school_names)):
        school_id, gaokao_school_name, match_type, confidence, notes = match_school_name(
            school_name,
            exact_lookup,
            normalized_lookup,
        )
        alias_from = MANUAL_SCHOOL_ALIASES.get(school_name) if match_type == "manual_alias" else None
        rows.append(
            (
                school_name,
                school_id,
                gaokao_school_name,
                match_type,
                confidence,
                alias_from,
                notes,
            )
        )
        result[school_name] = {
            "gaokao_school_id": school_id,
            "gaokao_school_name": gaokao_school_name,
            "match_type": match_type,
            "confidence": confidence,
            "notes": notes,
        }
    conn.executemany(
        """
        INSERT INTO school_mappings (
            school_name_official, gaokao_school_id, gaokao_school_name,
            match_type, confidence, alias_from, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return result


def fetch_plan_json(
    session: requests.Session,
    school_id: str,
    year: int,
) -> tuple[str, int, dict[str, Any] | None, str | None]:
    url = GAOKAO_SPECIAL_PLAN_URL.format(school_id=school_id, year=year)
    try:
        response = session.get(url, timeout=60)
        response.raise_for_status()
        payload = response.json()
        if payload.get("code") != "0000":
            return school_id, year, None, f"plan_code:{payload.get('code')}"
        return school_id, year, payload, None
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else "http_error"
        return school_id, year, None, f"http:{status_code}"
    except Exception as exc:  # noqa: BLE001
        return school_id, year, None, str(exc)


def dedupe_school_year_pairs(
    score_rows: list[ScoreRow],
    school_mapping: dict[str, dict[str, Any]],
) -> list[tuple[str, str, int]]:
    pairs = {
        (
            school_mapping[row.school_name]["gaokao_school_id"],
            school_mapping[row.school_name]["gaokao_school_name"],
            row.year,
        )
        for row in score_rows
        if school_mapping[row.school_name]["gaokao_school_id"]
    }
    return sorted((school_id, school_name, year) for school_id, school_name, year in pairs if school_id)


def insert_plan_majors(
    conn: sqlite3.Connection,
    plans: dict[tuple[str, int], dict[str, Any]],
    school_mapping: dict[str, dict[str, Any]],
) -> dict[tuple[int, str, str, str, str], list[int]]:
    group_index: dict[tuple[int, str, str, str, str], list[int]] = {}
    rows_to_insert: list[tuple[Any, ...]] = []
    seen_rows: set[tuple[Any, ...]] = set()

    for (school_id, year), payload in plans.items():
        data = payload.get("data", {})
        for key, value in data.items():
            if key in {"md5", "time"}:
                continue
            if not isinstance(value, dict) or "item" not in value:
                continue
            for item in value.get("item", []):
                if str(item.get("province", "")) != PROVINCE_ID:
                    continue
                local_batch_name = clean_text(item.get("local_batch_name"))
                if PLAN_LOCAL_BATCH_NAMES and local_batch_name not in PLAN_LOCAL_BATCH_NAMES:
                    continue
                type_code = clean_text(item.get("type"))
                if PLAN_TYPE_CODES and type_code not in PLAN_TYPE_CODES:
                    continue

                group_name = clean_text(item.get("sg_name"))
                group_digits = extract_digits(group_name)
                row_signature = (
                    year,
                    school_id,
                    PROVINCE_ID,
                    type_code,
                    clean_text(item.get("batch")),
                    clean_text(item.get("special_group")),
                    group_name,
                    clean_text(item.get("zslx_name")),
                    clean_text(item.get("special_id")),
                    clean_text(item.get("spcode")),
                    clean_text(item.get("spname")),
                    clean_text(item.get("tuition")),
                    numeric_int(item.get("num")),
                )
                if row_signature in seen_rows:
                    continue
                seen_rows.add(row_signature)
                row = (
                    year,
                    school_id,
                    next(
                        (
                            mapping["gaokao_school_name"]
                            for mapping in school_mapping.values()
                            if mapping["gaokao_school_id"] == school_id
                        ),
                        "",
                    ),
                    PROVINCE_ID,
                    type_code,
                    clean_text(item.get("batch")),
                    local_batch_name,
                    clean_text(item.get("special_group")),
                    group_name,
                    group_digits,
                    clean_text(item.get("sg_info")),
                    clean_text(item.get("sg_xuanke")),
                    clean_text(item.get("zslx_name")),
                    clean_text(item.get("zslx")),
                    clean_text(item.get("special_id")),
                    clean_text(item.get("spcode")),
                    clean_text(item.get("spname")),
                    clean_text(item.get("sp_name")),
                    clean_text(item.get("level1_name")),
                    clean_text(item.get("level2_name")),
                    clean_text(item.get("level3_name")),
                    clean_text(item.get("length")),
                    clean_text(item.get("tuition")),
                    numeric_int(item.get("num")),
                    clean_text(item.get("remark")),
                    clean_text(item.get("info")),
                    GAOKAO_SPECIAL_PLAN_URL.format(school_id=school_id, year=year),
                    json.dumps(item, ensure_ascii=False),
                )
                rows_to_insert.append(row)

    conn.executemany(
        """
        INSERT INTO plan_major_details (
            year, gaokao_school_id, gaokao_school_name, province_id, type_code, batch_code,
            local_batch_name, special_group_internal_id, special_group_display_name, group_digits,
            group_subject_info, group_subject_code, zslx_name, zslx_code, major_id, major_code,
            major_name, major_name_short, major_category_level1, major_category_level2,
            major_category_level3, study_length, tuition, enrollment_count, remark, extra_info,
            source_url, raw_payload
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows_to_insert,
    )

    cursor = conn.execute(
        """
        SELECT plan_major_id, year, gaokao_school_id, group_digits, COALESCE(zslx_name, ''), source_url
        FROM plan_major_details
        """
    )
    for plan_major_id, year, school_id, group_digits, zslx_name, source_url in cursor.fetchall():
        key = (year, school_id, group_digits, zslx_name, source_url)
        group_index.setdefault(key, []).append(plan_major_id)
    return group_index


def elective_info_matches_user(elective_info: str) -> bool:
    text = clean_text(elective_info)
    if not text:
        return False
    if "首选历史" in text:
        return False
    if "不限" in text:
        return True
    if "首选物理" not in text and "物理" not in text:
        return False

    mentioned = [subject for subject in ("化学", "生物", "政治", "地理") if subject in text]
    secondary = [subject for subject in mentioned if subject != "物理"]
    if not secondary:
        return True

    secondary_set = set(secondary)
    if any(marker in text for marker in ("2科必选", "均须选考", "都必须", "均应选考")):
        return secondary_set.issubset(USER_SUBJECTS)
    if any(marker in text for marker in ("2选1", "3选1", "1科必选", "任选", "选考其中一门", "选一门")):
        return bool(secondary_set & USER_SUBJECTS)
    if "/" in text or "或" in text:
        return bool(secondary_set & USER_SUBJECTS)
    if "化学、生物" in text or "化学、 生物" in text:
        return {"化学", "生物"}.issubset(USER_SUBJECTS)
    return secondary_set.issubset(USER_SUBJECTS)


def fetch_plan_major_payload(conn: sqlite3.Connection, plan_major_id: int) -> sqlite3.Row:
    cursor = conn.execute(
        """
        SELECT *
        FROM plan_major_details
        WHERE plan_major_id = ?
        """,
        (plan_major_id,),
    )
    row = cursor.fetchone()
    if row is None:
        raise KeyError(f"plan_major_id not found: {plan_major_id}")
    return row


def build_group_digit_candidates(score_row: ScoreRow) -> set[str]:
    code_digits = extract_digits(score_row.group_code)
    name_digits = extract_digits(score_row.group_name)
    candidates = {value for value in (code_digits, name_digits) if value}
    if ACTIVE_PROVINCE_SLUG != "jiangxi":
        return candidates

    if score_row.subject_type == "\u7269\u7406\u7c7b":
        for target_digits in list(candidates):
            if target_digits.isdigit():
                number = int(target_digits)
                if 500 <= number < 600:
                    candidates.add(str(number - 400))
    return {value for value in candidates if value}


def match_admission_groups_to_majors(
    conn: sqlite3.Connection,
    score_rows: list[ScoreRow],
    admission_mapping: dict[tuple[int, str, str, str], int],
    school_mapping: dict[str, dict[str, Any]],
) -> tuple[int, int]:
    conn.row_factory = sqlite3.Row
    eligible_rows: list[tuple[Any, ...]] = []
    unmatched_rows: list[tuple[Any, ...]] = []
    eligible_seen: set[tuple[Any, ...]] = set()

    plan_rows_by_school_year: dict[tuple[int, str], list[sqlite3.Row]] = {}
    cursor = conn.execute(
        """
        SELECT *
        FROM plan_major_details
        """
    )
    for row in cursor.fetchall():
        plan_rows_by_school_year.setdefault((row["year"], row["gaokao_school_id"]), []).append(row)

    for score_row in score_rows:
        admission_group_id = admission_mapping[(score_row.year, score_row.school_name, score_row.group_code, score_row.group_name)]
        school_info = school_mapping[score_row.school_name]
        school_id = school_info["gaokao_school_id"]
        if not school_id:
            unmatched_rows.append(
                (
                    admission_group_id,
                    score_row.year,
                    score_row.school_name,
                    score_row.group_code,
                    score_row.group_name,
                    "school_unmatched",
                    school_info["notes"],
                )
            )
            continue

        plan_rows = plan_rows_by_school_year.get((score_row.year, school_id), [])
        if not plan_rows:
            unmatched_rows.append(
                (
                    admission_group_id,
                    score_row.year,
                    score_row.school_name,
                    score_row.group_code,
                    score_row.group_name,
                    "plan_missing",
                    GAOKAO_SPECIAL_PLAN_URL.format(school_id=school_id, year=score_row.year),
                )
            )
            continue

        target_digit_candidates = build_group_digit_candidates(score_row)
        category = official_group_category(score_row.group_name)
        candidates = [
            row
            for row in plan_rows
            if clean_text(row["group_digits"]) in target_digit_candidates
        ]

        if not candidates:
            unmatched_rows.append(
                (
                    admission_group_id,
                    score_row.year,
                    score_row.school_name,
                    score_row.group_code,
                    score_row.group_name,
                    "group_not_found",
                    f"group_digits={','.join(sorted(target_digit_candidates))}; category={category}",
                )
            )
            continue

        unique_zslx = {clean_text(row["zslx_name"]) for row in candidates if clean_text(row["zslx_name"])}
        if len(unique_zslx) > 1:
            if category == "国家专项":
                candidates = [row for row in candidates if row["zslx_name"] == "国家专项"]
            elif category == "普通类":
                candidates = [row for row in candidates if row["zslx_name"] in {"", "普通类"}]
            elif category == "中外合作办学或国际项目":
                candidates = [
                    row
                    for row in candidates
                    if any(token in (clean_text(row["major_name"]) + clean_text(row["extra_info"]) + clean_text(row["remark"])) for token in ("中外合作", "国际项目"))
                ]

        if not candidates:
            unmatched_rows.append(
                (
                    admission_group_id,
                    score_row.year,
                    score_row.school_name,
                    score_row.group_code,
                    score_row.group_name,
                    "category_mismatch",
                    f"group_digits={','.join(sorted(target_digit_candidates))}; category={category}",
                )
            )
            continue

        eligible_candidates = [row for row in candidates if elective_info_matches_user(row["group_subject_info"])]
        if not eligible_candidates:
            unmatched_rows.append(
                (
                    admission_group_id,
                    score_row.year,
                    score_row.school_name,
                    score_row.group_code,
                    score_row.group_name,
                    "subject_ineligible",
                    next(iter({clean_text(row['group_subject_info']) for row in candidates}), ""),
                )
            )
            continue

        for row in eligible_candidates:
            eligible_signature = (
                admission_group_id,
                clean_text(row["major_code"]),
                clean_text(row["major_name"]),
                clean_text(row["tuition"]),
                clean_text(row["study_length"]),
                clean_text(row["zslx_name"]),
            )
            if eligible_signature in eligible_seen:
                continue
            eligible_seen.add(eligible_signature)
            eligible_rows.append(
                (
                    admission_group_id,
                    row["plan_major_id"],
                    score_row.year,
                    score_row.school_code,
                    score_row.school_name,
                    score_row.group_code,
                    score_row.group_name,
                    category,
                    score_row.score,
                    row["group_subject_info"],
                    row["major_code"],
                    row["major_name"],
                    row["major_name_short"],
                    row["major_category_level1"],
                    row["major_category_level2"],
                    row["major_category_level3"],
                    row["study_length"],
                    row["tuition"],
                    row["enrollment_count"],
                    row["zslx_name"],
                    row["source_url"],
                )
            )

    conn.executemany(
        """
        INSERT INTO eligible_majors (
            admission_group_id, plan_major_id, year, school_code, school_name,
            official_group_code, official_group_name, official_group_category,
            score, elective_info, major_code, major_name, major_name_short,
            major_category_level1, major_category_level2, major_category_level3,
            study_length, tuition, enrollment_count, zslx_name, source_url
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        eligible_rows,
    )
    conn.executemany(
        """
        INSERT INTO unmatched_groups (
            admission_group_id, year, school_name, official_group_code,
            official_group_name, reason, detail
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        unmatched_rows,
    )
    return len(eligible_rows), len(unmatched_rows)


def insert_run_summary(
    conn: sqlite3.Connection,
    started_at: str,
    finished_at: str,
    score_group_count: int,
    school_count: int,
    matched_school_count: int,
    plan_fetch_success_count: int,
    plan_fetch_failure_count: int,
    eligible_major_count: int,
    unmatched_group_count: int,
) -> None:
    subject_label = "+".join(sorted(USER_SUBJECTS))
    conn.execute(
        """
        INSERT INTO crawl_runs (
            started_at, finished_at, score_group_count, school_count, matched_school_count,
            plan_fetch_success_count, plan_fetch_failure_count, eligible_major_count,
            unmatched_group_count, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            started_at,
            finished_at,
            score_group_count,
            school_count,
            matched_school_count,
            plan_fetch_success_count,
            plan_fetch_failure_count,
            eligible_major_count,
            unmatched_group_count,
            f"可报口径：用户选科为{subject_label}，按专业组选科要求可满足即视为可报。",
        ),
    )


def finalize_database_path(temp_db_path: Path) -> Path:
    if temp_db_path == DB_PATH:
        return DB_PATH
    try:
        if DB_PATH.exists():
            DB_PATH.unlink()
        temp_db_path.replace(DB_PATH)
        return DB_PATH
    except PermissionError:
        fallback_path = DB_PATH.with_name(
            f"{DB_PATH.stem}_rebuilt_{datetime.now().strftime('%Y%m%d_%H%M%S')}{DB_PATH.suffix}"
        )
        if fallback_path.exists():
            fallback_path.unlink()
        temp_db_path.replace(fallback_path)
        return fallback_path


def build_database() -> None:
    started_at = now_iso()
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            "Accept": "application/json,text/plain,*/*",
        }
    )

    extract_dir = ensure_raw_sources(session)
    ensure_gaokao_school_list(session)
    score_rows = load_filtered_scores(extract_dir)
    legacy_school_line_rows = load_legacy_school_line_rows()
    gaokao_school_data = load_gaokao_school_map()
    exact_lookup, normalized_lookup = build_school_lookup(gaokao_school_data)

    temp_db_path = DB_PATH.with_name(f"{DB_PATH.stem}.build{DB_PATH.suffix}")
    if temp_db_path.exists():
        temp_db_path.unlink()
    conn = sqlite3.connect(temp_db_path)
    create_schema(conn)
    insert_sources(conn)
    insert_legacy_school_lines(conn, legacy_school_line_rows)
    admission_mapping = insert_admission_groups(conn, score_rows)
    school_mapping = insert_school_mappings(
        conn,
        [row.school_name for row in score_rows],
        exact_lookup,
        normalized_lookup,
    )
    conn.commit()

    school_year_pairs = dedupe_school_year_pairs(score_rows, school_mapping)
    matched_school_count = len({row.school_name for row in score_rows if school_mapping[row.school_name]["gaokao_school_id"]})
    school_count = len(set(row.school_name for row in score_rows))

    plans: dict[tuple[str, int], dict[str, Any]] = {}
    failures: list[tuple[str, int, str | None]] = []
    lock = threading.Lock()

    def worker(args: tuple[str, str, int]) -> None:
        school_id, _school_name, year = args
        local_session = requests.Session()
        local_session.headers.update(session.headers)
        result_school_id, result_year, payload, error = fetch_plan_json(local_session, school_id, year)
        with lock:
            if payload is not None:
                plans[(result_school_id, result_year)] = payload
            else:
                failures.append((result_school_id, result_year, error))

    with concurrent.futures.ThreadPoolExecutor(max_workers=12) as executor:
        list(executor.map(worker, school_year_pairs))

    insert_plan_majors(conn, plans, school_mapping)
    conn.commit()

    eligible_major_count, unmatched_group_count = match_admission_groups_to_majors(
        conn,
        score_rows,
        admission_mapping,
        school_mapping,
    )

    finished_at = now_iso()
    insert_run_summary(
        conn,
        started_at=started_at,
        finished_at=finished_at,
        score_group_count=len(score_rows),
        school_count=school_count,
        matched_school_count=matched_school_count,
        plan_fetch_success_count=len(plans),
        plan_fetch_failure_count=len(failures),
        eligible_major_count=eligible_major_count,
        unmatched_group_count=unmatched_group_count,
    )
    conn.commit()
    conn.close()

    final_db_path = finalize_database_path(temp_db_path)
    print(f"province: {ACTIVE_PROVINCE_SLUG}")
    print(f"database: {final_db_path}")
    print(f"score_groups: {len(score_rows)}")
    print(f"matched_schools: {matched_school_count}/{school_count}")
    print(f"plan_fetch_success: {len(plans)}")
    print(f"plan_fetch_failure: {len(failures)}")
    print(f"eligible_majors: {eligible_major_count}")
    print(f"unmatched_groups: {unmatched_group_count}")

def main() -> None:
    parser = argparse.ArgumentParser(description="Build the admissions SQLite database.")
    parser.add_argument("--province", default=DEFAULT_PROVINCE_SLUG)
    args = parser.parse_args()

    configure_runtime(args.province)
    validate_runtime_build_ready()
    build_database()


if __name__ == "__main__":
    main()
